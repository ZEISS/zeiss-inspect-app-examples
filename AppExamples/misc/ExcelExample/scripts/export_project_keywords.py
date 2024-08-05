# -*- coding: utf-8 -*-
#
# export_project_keywords.py
#
# Carl Zeiss GOM Metrology GmbH, 2024
#
# This App is part of the ZEISS INSPECT Python API Examples:
# https://zeissiqs.github.io/zeiss-inspect-addon-api/2025/python_examples/
# ---

import gom
from openpyxl import Workbook
from openpyxl.styles import Font
import datetime

if not hasattr(gom.app, 'project'):
    gom.script.sys.execute_user_defined_dialog(file='no_project.gdlg')
    quit(0)

# Repeat until file was written successfully or
# user has cancelled.
while True:
    try:
        # File selection dialog
        RESULT1 = gom.script.sys.execute_user_defined_dialog(file='export_file.gdlg')
    except gom.BreakError:
        # User has cancelled
        break

    # Create a workbook
    wb = Workbook()

    # Select the active worksheet
    ws = wb.active

    # Create table header
    ws['A1'] = "Project Keyword"
    ws['B1'] = "Description"
    ws['C1'] = "Value"

    # Change table header layout
    for cell in ['A1', 'B1', 'C1']:
        ws[cell].font = Font(bold=True, size=16)
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 70
    ws.column_dimensions['C'].width = 50

    for key in gom.app.project.project_keywords:
        val = gom.app.project.get(key)
        desc = gom.app.project.get(f'description({key})')

        # Remove prefix 'user_' from key
        key = key[5:]
        print(f"{key} - {desc}: {val}")

        # Special case - convert gom.Date to datetime-object
        # and format Excel cell accordingly
        if type(val) is gom.Date:
            val = datetime.datetime(val.year, val.month, val.day)
            ws.append([key, desc, val])
            ws.cell(row=ws.max_row, column=2).number_format = "yyyy-mm-dd"
        else:
            ws.append([key, desc, val])

    # Empty row
    ws.append([])

    # Project file
    ws.append(["Project file:", gom.app.project.project_file])

    # Metadata
    ws.cell(row=ws.max_row + 2, column=1).value = f'Exported from {gom.app.application_name} \
	{gom.app.application_build_information.version} \
	Rev. {gom.app.application_build_information.revision} \
	by {gom.app.current_user} on {gom.app.current_date}'

    ws.cell(row=ws.max_row, column=1).font = Font(size=8)

    # Try to write Excel file, this will fail if the file is still open in Excel!
    retry = False
    try:
        wb.save(RESULT1.file)
    except PermissionError:
        retry = True
        # Notification dialog
        RESULT = gom.script.sys.execute_user_defined_dialog(dialog={
            "content": [
                [
                    {
                        "columns": 1,
                        "data": "AAAAAA==",
                                "file_name": "",
                                "height": 0,
                                "keep_aspect": True,
                                "keep_original_size": True,
                                "name": "image",
                                "rows": 1,
                                "system_image": "system_message_critical",
                                "tooltip": {
                                        "id": "",
                                        "text": "",
                                        "translatable": True
                                },
                        "type": "image",
                        "use_system_image": True,
                        "width": 0
                    },
                    {
                        "columns": 1,
                        "name": "label",
                                "rows": 1,
                                "text": {
                                    "id": "",
                                    "text": "File cannot be written!",
                                    "translatable": True
                                },
                        "tooltip": {
                                    "id": "",
                                    "text": "",
                                    "translatable": True
                                },
                        "type": "label",
                        "word_wrap": False
                    }
                ]
            ],
            "control": {
                "id": "Close"
            },
            "embedding": "always_toplevel",
            "position": "automatic",
            "size": {
                "height": 119,
                "width": 195
            },
            "sizemode": "automatic",
            "style": "",
            "title": {
                "id": "",
                "text": "Message",
                "translatable": True
            }
        })

    if not retry:
        break
